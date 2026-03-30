#!/usr/bin/env python3
"""
Import LOT inward Excel sheets into live inventory using the same
manual API endpoint: POST /api/inventory/inward.

Example:
  python3 scripts/import_lot_inward_xlsx.py \
    /Users/deepakios/Desktop/LOT\ INWARD.xlsx \
    --api-base http://localhost:5001/api \
    --email admin@example.com --password secret

Or with token:
  python3 scripts/import_lot_inward_xlsx.py /path/file.xlsx --token <JWT>
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
import urllib.error
import urllib.request
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    print(f"Failed to import openpyxl: {exc}")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value)
    if not text or text in {"-", "--"}:
        return None

    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def format_weight(value: float) -> str:
    text = f"{value:.3f}".rstrip("0").rstrip(".")
    return text if text else "0"


def extract_prefixed_value(ws: Any, prefixes: List[str], max_scan_rows: int = 12) -> str:
    patterns = [re.compile(rf"^{re.escape(prefix)}\s*-\s*(.+)$", re.IGNORECASE) for prefix in prefixes]
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=row, column=col).value
            text = normalize_text(raw)
            if not text:
                continue
            for pattern in patterns:
                match = pattern.match(text)
                if match:
                    return normalize_text(match.group(1))
    return ""


def parse_inward_date(raw: str) -> str:
    text = normalize_text(raw)
    if not text:
        return dt.date.today().isoformat()

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))
        if year < 100:
            year += 2000
        return dt.date(year, month, day).isoformat()

    return dt.date.today().isoformat()


def build_sheet_payload(ws: Any) -> Optional[Dict[str, Any]]:
    lot_no = extract_prefixed_value(ws, ["LOT NO", "LOT NUMBER"])
    inward_date_raw = extract_prefixed_value(ws, ["DATE"])
    from_party = extract_prefixed_value(ws, ["PARTY NAME", "PARTY"])
    lot_name = extract_prefixed_value(ws, ["LOT NAME"])
    dia = extract_prefixed_value(ws, ["DIA"])

    if not lot_no or not lot_name or not from_party or not dia:
        return None

    # Expected layout in provided file:
    # row 5 -> rack labels per set column
    # row 6 -> pallet labels per set column
    # row 7 -> COLOUR | S-1...S-N | TOTAL
    # row 8+ -> colour rows
    header_row = 7
    set_cols: List[int] = []

    for col in range(2, ws.max_column + 1):
        value = normalize_text(ws.cell(row=header_row, column=col).value)
        if not value:
            if set_cols:
                break
            continue
        upper = value.upper()
        if upper.startswith("TOTAL"):
            break
        if upper.startswith("S-") or upper.startswith("SET"):
            set_cols.append(col)

    if not set_cols:
        return None

    set_labels = [
        normalize_text(ws.cell(row=header_row, column=col).value) or f"S-{idx + 1}"
        for idx, col in enumerate(set_cols)
    ]
    racks = [normalize_text(ws.cell(row=5, column=col).value) for col in set_cols]
    pallets = [normalize_text(ws.cell(row=6, column=col).value) for col in set_cols]

    rows: List[Dict[str, Any]] = []
    rec_roll = 0
    rec_wt = 0.0
    started = False

    for row in range(8, ws.max_row + 1):
        colour = normalize_text(ws.cell(row=row, column=1).value)
        if not colour:
            if started:
                break
            continue
        started = True

        set_weights: List[str] = []
        total_weight = 0.0
        row_rolls = 0

        for col in set_cols:
            weight = parse_number(ws.cell(row=row, column=col).value)
            if weight is None:
                set_weights.append("")
            else:
                set_weights.append(format_weight(weight))
                total_weight += weight
                row_rolls += 1

        if row_rolls == 0:
            continue

        rows.append(
            {
                "colour": colour,
                "gsm": "",
                "setWeights": set_weights,
                "setLabels": set_labels,
                "totalWeight": round(total_weight, 3),
            }
        )
        rec_roll += row_rolls
        rec_wt += total_weight

    if not rows:
        return None

    inward_date = parse_inward_date(inward_date_raw)
    process = "COMPACTING" if "COMPACT" in from_party.upper() else ""

    return {
        "inwardDate": inward_date,
        "inTime": "09:00 AM",
        "outTime": "09:30 AM",
        "lotName": lot_name,
        "lotNo": lot_no,
        "fromParty": from_party,
        "process": process,
        "rate": 0,
        "gsm": "",
        "vehicleNo": "",
        "partyDcNo": "",
        "diaEntries": [
            {
                "dia": dia,
                "roll": rec_roll,
                "sets": len(set_cols),
                "delivWt": round(rec_wt, 3),
                "recRoll": rec_roll,
                "recWt": round(rec_wt, 3),
                "rate": 0,
            }
        ],
        "storageDetails": [
            {
                "dia": dia,
                "racks": racks,
                "pallets": pallets,
                "rows": rows,
            }
        ],
        "qualityStatus": "OK",
        "gsmStatus": "OK",
        "shadeStatus": "OK",
        "washingStatus": "OK",
        "complaintText": "",
    }


def post_json(url: str, payload: Dict[str, Any], token: Optional[str], timeout: int = 45) -> Dict[str, Any]:
    body = json.dumps(payload).encode("utf-8")
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"

    req = urllib.request.Request(url=url, data=body, headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        text = resp.read().decode("utf-8")
        if not text:
            return {}
        return json.loads(text)


def login(api_base: str, email: str, password: str, timeout: int = 45) -> str:
    url = f"{api_base.rstrip('/')}/auth/login"
    try:
        response = post_json(url, {"email": email, "password": password}, token=None, timeout=timeout)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code}: {body}") from exc
    token = response.get("token")
    if not token:
        raise RuntimeError("Login succeeded but token was missing in response")
    return token


def main() -> int:
    parser = argparse.ArgumentParser(description="Bulk import LOT inward Excel into live inventory")
    parser.add_argument("xlsx_path", help="Path to LOT INWARD .xlsx file")
    parser.add_argument("--api-base", default="http://localhost:5001/api", help="API base URL")
    parser.add_argument("--token", default=os.getenv("GARMENTS_TOKEN"), help="Bearer token")
    parser.add_argument("--email", default=os.getenv("GARMENTS_EMAIL"), help="Login email (if token not set)")
    parser.add_argument("--password", default=os.getenv("GARMENTS_PASSWORD"), help="Login password (if token not set)")
    parser.add_argument("--dry-run", action="store_true", help="Parse only; do not upload")
    parser.add_argument("--stop-on-error", action="store_true", help="Stop on first failed sheet")
    parser.add_argument("--timeout", type=int, default=45, help="HTTP timeout seconds")
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    entries: List[Dict[str, Any]] = []
    skipped: List[str] = []

    for ws in workbook.worksheets:
        payload = build_sheet_payload(ws)
        if payload is None:
            skipped.append(ws.title)
            continue
        payload["_sheet"] = ws.title
        entries.append(payload)

    print(f"Workbook: {args.xlsx_path}")
    print(f"Sheets parsed: {len(entries)}")
    if skipped:
        print(f"Sheets skipped: {', '.join(skipped)}")

    for entry in entries:
        sheet = entry["_sheet"]
        dia_entry = entry["diaEntries"][0]
        print(
            f"- {sheet}: lot={entry['lotNo']}, date={entry['inwardDate']}, dia={dia_entry['dia']}, "
            f"rows={len(entry['storageDetails'][0]['rows'])}, recWt={dia_entry['recWt']}"
        )

    if args.dry_run:
        print("Dry-run complete. No upload performed.")
        return 0

    token = args.token
    if not token:
        if not args.email or not args.password:
            print("Provide --token or both --email and --password")
            return 2
        try:
            token = login(args.api_base, args.email, args.password, timeout=args.timeout)
            print("Login successful.")
        except Exception as exc:
            print(f"Login failed: {exc}")
            return 2

    success_count = 0
    failure_count = 0

    inward_url = f"{args.api_base.rstrip('/')}/inventory/inward"

    for entry in entries:
        sheet = entry.pop("_sheet", "Unknown")
        try:
            post_json(inward_url, entry, token=token, timeout=args.timeout)
            success_count += 1
            print(f"[OK] {sheet} -> {entry['lotNo']}")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            failure_count += 1
            print(f"[FAIL] {sheet} -> {entry['lotNo']} | HTTP {exc.code} | {body}")
            if args.stop_on_error:
                break
        except Exception as exc:
            failure_count += 1
            print(f"[FAIL] {sheet} -> {entry['lotNo']} | {exc}")
            if args.stop_on_error:
                break

    print("\nImport summary")
    print(f"Success: {success_count}")
    print(f"Failed : {failure_count}")

    return 0 if failure_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
